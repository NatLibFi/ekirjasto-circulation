"""
This module provides tools for exporting analytics data from the system
into CSV or Excel formats.
"""

from io import BytesIO
from tempfile import NamedTemporaryFile

import unicodecsv as csv
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from sqlalchemy import Integer
from sqlalchemy.sql import cast, func, select
from sqlalchemy.sql.expression import and_, case, join, literal_column, or_

from core.model import (
    CirculationEvent,
    Collection,
    DataSource,
    Edition,
    Genre,
    Identifier,
    IntegrationConfiguration,
    Library,
    LicensePool,
    Work,
    WorkGenre,
)

# Finland:
from core.model.contributor import Contribution, Contributor


class LocalAnalyticsExporter:
    """Export large numbers of analytics events in CSV or Excel format."""

    def _fetch_and_process_data(self, _db, start, end, locations, library):
        """Fetches data, converts them into a list of dicts and calculates max contributors.

        Args:
            _db: The database session.
            start: The start date for the analytics data.
            end: The end date for the analytics data.
            locations: A comma-separated string of locations to filter by.
            library: The library to filter by.

        Returns:
            A tuple containing:
                - A list of dictionaries, where each dictionary represents a row of data.
                - The maximum number of contributors found in any row.
        """
        query = self.analytics_query_loan_statistics(start, end, locations, library)
        results = _db.execute(query)
        rows = [dict(row) for row in results]

        max_contributors = 0
        for row in rows:
            contributors = row.get("contributors", []) or []
            max_contributors = max(max_contributors, len(contributors))

        return rows, max_contributors

    def export(self, _db, start, end, locations=None, library=None):
        """Exports analytics data into a CSV string.

        Args:
            _db: The database session.
            start: The start date for the analytics data.
            end: The end date for the analytics data.
            locations: A comma-separated string of locations to filter by.
            library: The library to filter by.

        Returns:
            A string containing the CSV data.
        """
        query = self.analytics_query(start, end, locations, library)
        results = _db.execute(query)

        # Write the CSV file to a BytesIO.
        header = [
            "time",
            "event",
            "identifier",
            "identifier_type",
            "title",
            "author",
            "fiction",
            "audience",
            "publisher",
            "imprint",
            "language",
            "target_age",
            "genres",
            "location",
            "collection_name",
            "library_short_name",
            "library_name",
            "medium",
            "distributor",
            "open_access",
        ]
        output = BytesIO()
        writer = csv.writer(output, encoding="utf-8")
        writer.writerow(header)
        writer.writerows(results)
        return output.getvalue().decode("utf-8")

    def export_excel(self, _db, start, end, locations=None, library=None):
        """Exports loan statistics data into an Excel file.

        Args:
            _db: The database session.
            start: The start date for the analytics data.
            end: The end date for the analytics data.
            locations: A comma-separated string of locations to filter by.
            library: The library to filter by.

        Returns:
            A bytes object containing the Excel file data.
        """
        rows, max_contributors = self._fetch_and_process_data(
            _db, start, end, locations, library
        )

        # Prepare Excel workbook
        workbook = Workbook()
        sheet = workbook.active

        header = [
            "Tekijä (aakkostus)",
            "Nimeke",
            "Lajityyppi",
            "Tunniste",
            "Tunnisteen tyyppi",
            "Kirjasto",
            "Sijainti",
            "Formaatti",
            "Kategoria(t)",
            "Kohderyhmä",
            "Kieli",
            "Kustantaja/Julkaisija",
            "Julkaisuvuosi",
            "Kaikki lainat",
        ]
        # Add headers for contributor columns.
        # These will be populated with all contributor names (authors and others).
        for i in range(max_contributors):
            header.append(f"Tekijä {i+1}")

        sheet.append(header)

        for row in rows:
            genres = row.get("genres")
            categories = ", ".join(genres) if genres else ""

            contributors = row.get("contributors") or []

            sheet.append(
                [
                    # Tekijä (aakkostus)
                    row.get("sort_author", ""),
                    # Nimeke
                    row.get("sort_title", ""),
                    # Fiktio
                    "kaunokirjallisuus" if row.get("fiction") else "tietokirjallisuus",
                    # Tunniste
                    row.get("identifier", ""),
                    # Tunnisteen tyyppi
                    row.get("identifier_type", ""),
                    # Kirjasto
                    row.get("library_name", ""),
                    # Sijainti
                    row.get("location", ""),
                    # Formaatti
                    row.get("medium", ""),
                    # Kategoria(t)
                    categories,
                    # Kohderyhmä
                    row.get("audience", ""),
                    # Kieli
                    row.get("language", ""),
                    # Kustantaja/Julkaisija
                    row.get("publisher", ""),
                    # Julkaisuvuosi
                    row.get("published_year"),
                    # Kaikki lainat
                    row.get("count", ""),
                    # Tekijät (all contributors) (1-n rows)
                    *contributors,  # Using combined 'contributors' list
                ]
            )

        ### Adjust styles
        column_width = 24

        # Loop through all columns and set the width
        for column in sheet.columns:
            for cell in column:
                sheet.column_dimensions[cell.column_letter].width = column_width

        # Define styles for the header row
        header_style = Font(name="Calibri", bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="336699", end_color="336699", fill_type="solid"
        )
        header_border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Apply styles to the header row
        for cell in sheet[1]:
            cell.font = header_style
            cell.fill = header_fill
            cell.border = header_border

        # Make header row sticky
        sheet.freeze_panes = "A2"

        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        return stream

    def export_csv(self, _db, start, end, locations=None, library=None):
        """Exports loan statistics data into a CSV file.

        Args:
            _db: The database session.
            start: The start date for the analytics data.
            end: The end date for the analytics data.
            locations: A comma-separated string of locations to filter by.
            library: The library to filter by.

        Returns:
            A string containing the CSV data.
        """
        rows, max_contributors = self._fetch_and_process_data(
            _db, start, end, locations, library
        )

        # Prepare CSV output
        header = [
            "Tekijä (aakkostus)",
            "Nimeke",
            "Lajityyppi",
            "Tunniste",
            "Tunnisteen tyyppi",
            "Kirjasto",
            "Sijainti",
            "Formaatti",
            "Kategoria(t)",
            "Kohderyhmä",
            "Kieli",
            "Kustantaja/Julkaisija",
            "Julkaisuvuosi",
            "Kaikki lainat",
        ]

        # Add headers for contributor columns.
        # These will be populated with all contributor names (authors and others).
        for i in range(max_contributors):
            header.append(f"Tekijä {i+1}")

        output = BytesIO()
        writer = csv.writer(output, encoding="utf-8")
        writer.writerow(header)

        for row in rows:
            genres = row.get("genres")
            categories = ", ".join(genres) if genres else ""

            contributors = row.get("contributors") or []

            writer.writerow(
                [
                    # Tekijä (aakkostus)
                    row.get("sort_author", ""),
                    # Nimeke
                    row.get("sort_title", ""),
                    # Fiktio
                    "kaunokirjallisuus" if row.get("fiction") else "tietokirjallisuus",
                    # Tunniste
                    row.get("identifier", ""),
                    # Tunnisteen tyyppi
                    row.get("identifier_type", ""),
                    # Kirjasto
                    row.get("library_name", ""),
                    # Sijainti
                    row.get("location", ""),
                    # Formaatti
                    row.get("medium", ""),
                    # Kategoria(t)
                    categories,
                    # Kohderyhmä
                    row.get("audience", ""),
                    # Kieli
                    row.get("language", ""),
                    # Kustantaja/Julkaisija
                    row.get("publisher", ""),
                    # Julkaisuvuosi
                    row.get("published_year"),
                    # Kaikki lainat
                    row.get("count", ""),
                    # Tekijät (all contributors) (1-n rows)
                    *contributors,  # Using combined 'contributors' list
                ]
            )
        return output.getvalue().decode("utf-8")

    def analytics_query(self, start, end, locations=None, library=None):
        """Build a database query that fetches rows of analytics data.

        This method uses low-level SQLAlchemy code to do all
        calculations and data conversations in the database. It's
        modeled after Work.to_search_documents, which generates a
        large JSON document entirely in the database.

        Args:
            start: The start date for the analytics data.
            end: The end date for the analytics data.
            locations: A comma-separated string of locations to filter by.
            library: The library to filter by.

        Returns:
            A SQLAlchemy Select object representing the query.
        """

        clauses = [
            CirculationEvent.start >= start,
            CirculationEvent.start < end,
        ]

        if locations:
            event_types = [
                CirculationEvent.CM_CHECKOUT,
                CirculationEvent.CM_FULFILL,
                CirculationEvent.OPEN_BOOK,
            ]
            locations = locations.strip().split(",")

            clauses += [
                CirculationEvent.type.in_(event_types),
                CirculationEvent.location.in_(locations),
            ]

        if library:
            clauses += [CirculationEvent.library == library]

        # Build the primary query. This is a query against the
        # CirculationEvent table and a few other tables joined against
        # it. This makes up the bulk of the data.
        events_alias = (
            select(
                [
                    func.to_char(CirculationEvent.start, "YYYY-MM-DD HH24:MI:SS").label(
                        "start"
                    ),
                    CirculationEvent.type.label("event_type"),
                    Identifier.identifier,
                    Identifier.type.label("identifier_type"),
                    Edition.sort_title,
                    Edition.sort_author,
                    case(
                        [(Work.fiction == True, literal_column("'fiction'"))],
                        else_=literal_column("'nonfiction'"),
                    ).label("fiction"),
                    Work.id.label("work_id"),
                    Work.audience,
                    Edition.publisher,
                    Edition.imprint,
                    Edition.language,
                    CirculationEvent.location,
                    IntegrationConfiguration.name.label("collection_name"),
                    Library.short_name.label("library_short_name"),
                    Library.name.label("library_name"),
                    Edition.medium,
                    DataSource.name.label("distributor"),
                    LicensePool.open_access,
                ],
            )
            .select_from(
                join(
                    CirculationEvent,
                    LicensePool,
                    CirculationEvent.license_pool_id == LicensePool.id,
                )
                .join(Identifier, LicensePool.identifier_id == Identifier.id)
                .join(Work, Work.id == LicensePool.work_id)
                .join(Edition, Work.presentation_edition_id == Edition.id)
                .join(Collection, LicensePool.collection_id == Collection.id)
                .join(
                    IntegrationConfiguration,
                    Collection.integration_configuration_id
                    == IntegrationConfiguration.id,
                )
                .join(DataSource, LicensePool.data_source_id == DataSource.id)
                .outerjoin(Library, CirculationEvent.library_id == Library.id)
            )
            .where(and_(*clauses))
            .order_by(CirculationEvent.start.asc())
            .alias("events_alias")
        )

        # A subquery can hook into the main query by referencing its
        # 'work_id' field in its WHERE clause.
        work_id_column = literal_column(
            events_alias.name + "." + events_alias.c.work_id.name
        )

        # This subquery gets the names of a Work's genres as a single
        # comma-separated string.
        #

        # This Alias selects some number of rows, each containing one
        # string column (Genre.name). Genres with higher affinities with
        # this work go first.
        genres_alias = (
            select([Genre.name.label("genre_name")])
            .select_from(join(WorkGenre, Genre, WorkGenre.genre_id == Genre.id))
            .where(WorkGenre.work_id == work_id_column)
            .order_by(WorkGenre.affinity.desc(), Genre.name)
            .alias("genres_subquery")
        )

        # Use array_agg() to consolidate the rows into one row -- this
        # gives us a single value, an array of strings, for each
        # Work. Then use array_to_string to convert the array into a
        # single comma-separated string.
        genres = select(
            [func.array_to_string(func.array_agg(genres_alias.c.genre_name), ",")]
        ).select_from(genres_alias)

        # This subquery gets the a Work's target age as a single string.
        #

        # This Alias selects two fields: the lower and upper bounds of
        # the Work's target age. This reuses code originally written
        # for Work.to_search_documents().
        target_age = Work.target_age_query(work_id_column).alias("target_age_subquery")

        # Concatenate the lower and upper bounds with a dash in the
        # middle. If both lower and upper bound are empty, just give
        # the empty string. This simulates the behavior of
        # Work.target_age_string.
        target_age_string = select(
            [
                case(
                    [
                        (
                            or_(target_age.c.lower != None, target_age.c.upper != None),
                            func.concat(target_age.c.lower, "-", target_age.c.upper),
                        )
                    ],
                    else_=literal_column("''"),
                )
            ]
        ).select_from(target_age)

        # Build the main query out of the subqueries.
        events = events_alias.c
        query = select(
            [
                events.start,
                events.event_type,
                events.identifier,
                events.identifier_type,
                events.sort_title,
                events.sort_author,
                events.fiction,
                events.audience,
                events.publisher,
                events.imprint,
                events.language,
                target_age_string.label("target_age"),
                genres.label("genres"),
                events.location,
                events.collection_name,
                events.library_short_name,
                events.library_name,
                events.medium,
                events.distributor,
                case({True: "true", False: "false"}, value=events.open_access),
            ]
        ).select_from(events_alias)
        return query

    # Finland
    def analytics_query_loan_statistics(self, start, end, locations=None, library=None):
        """Build a database query that fetches analytics data
        for loan statistics export.

        Heavily modified from analytics_query method.

        This method uses low-level SQLAlchemy code to do all
        calculations and data conversations in the database.

        Args:
            start: The start date for the analytics data.
            end: The end date for the analytics data.
            locations: A comma-separated string of locations to filter by.
            library: The library to filter by.

        Returns:
            A SQLAlchemy Select object representing the query.
        """

        clauses = []

        # Filter by date range
        if start:
            clauses += [CirculationEvent.start >= start]
        if end:
            clauses += [CirculationEvent.start < end]

        # Take only checkout events
        clauses += [
            CirculationEvent.type.in_(
                [CirculationEvent.CM_CHECKOUT, CirculationEvent.DISTRIBUTOR_CHECKOUT]
            )
        ]

        if locations:
            locations = locations.strip().split(",")

            clauses += [
                CirculationEvent.location.in_(locations),
            ]

        if library:
            clauses += [CirculationEvent.library == library]

        # Build the primary query. This is a query against the
        # CirculationEvent table and a few other tables joined against
        # it. This makes up the bulk of the data.
        events_alias = (
            select(
                [
                    Identifier.identifier,
                    Identifier.type.label("identifier_type"),
                    Edition.sort_title,
                    Edition.sort_author,
                    case(
                        [(Work.fiction == True, True)],
                        else_=False,
                    ).label("fiction"),
                    Work.audience,
                    Work.id.label("work_id"),
                    Edition.publisher,
                    Edition.language,
                    CirculationEvent.location,
                    Library.name.label("library_name"),
                    Edition.medium,
                    Edition.id.label("edition_id"),
                    func.count().label("count"),
                    cast(func.extract("year", Edition.published), Integer).label(
                        "published_year"
                    ),
                ],
            )
            .select_from(
                join(
                    CirculationEvent,
                    LicensePool,
                    CirculationEvent.license_pool_id == LicensePool.id,
                )
                .join(Identifier, LicensePool.identifier_id == Identifier.id)
                .join(Work, Work.id == LicensePool.work_id)
                .join(Edition, Work.presentation_edition_id == Edition.id)
                .join(Collection, LicensePool.collection_id == Collection.id)
                .outerjoin(Library, CirculationEvent.library_id == Library.id)
            )
            .where(and_(*clauses))
            .group_by(
                Work.id,
                Identifier.identifier,
                Identifier.type.label("identifier_type"),
                Edition.sort_title,
                Edition.sort_author,
                Work.id.label("work_id"),
                Edition.publisher,
                Edition.language,
                CirculationEvent.location,
                Library.name.label("library_name"),
                Edition.id.label("edition_id"),
                Edition.medium,
                Edition.published,
            )
            .order_by(Edition.sort_author.asc())
            .alias("events_alias")
        )

        edition_id_column = literal_column(
            events_alias.name + "." + events_alias.c.edition_id.name
        )

        # Fetch all contributors (authors and others)
        contributors_alias = (
            select(
                [
                    Contributor.sort_name,
                    Contribution.role,
                ]
            )
            .where(
                Contribution.edition_id == edition_id_column,
            )
            .select_from(
                join(
                    Contributor,
                    Contribution,
                    Contributor.id == Contribution.contributor_id,
                )
            )
            .alias("contributors_alias")
        )

        # Combine contributor sort_name with role, eg. "sortname (role)" in a subquery
        contributors_subquery = select(
            [
                func.concat(
                    contributors_alias.c.sort_name, " (", contributors_alias.c.role, ")"
                ).label("contributor_with_role")
            ]
        ).select_from(contributors_alias)

        contributors = select(
            [
                func.array_agg(contributors_subquery.c.contributor_with_role).label(
                    "contributors_with_roles"
                )
            ]
        ).select_from(contributors_subquery)

        # A subquery can hook into the main query by referencing its
        # 'work_id' field in its WHERE clause.
        work_id_column = literal_column(
            events_alias.name + "." + events_alias.c.work_id.name
        )

        # This subquery gets the names of a Work's genres as a single
        # comma-separated string.
        #

        # This Alias selects some number of rows, each containing one
        # string column (Genre.name). Genres with higher affinities with
        # this work go first.
        genres_alias = (
            select([Genre.name.label("genre_name")])
            .select_from(join(WorkGenre, Genre, WorkGenre.genre_id == Genre.id))
            .where(WorkGenre.work_id == work_id_column)
            .order_by(WorkGenre.affinity.desc(), Genre.name)
            .alias("genres_subquery")
        )

        # Use array_agg() to consolidate the rows into one row -- this
        # gives us a single value, an array of strings, for each
        # Work.
        genres = select([func.array_agg(genres_alias.c.genre_name)]).select_from(
            genres_alias
        )

        # Build the main query out of the subqueries.
        events = events_alias.c
        query = select(
            [
                events.identifier,
                events.identifier_type,
                events.sort_title,
                events.sort_author,
                events.fiction,
                events.publisher,
                events.language,
                genres.label("genres"),
                events.audience,
                contributors.label("contributors"),
                events.location,
                events.library_name,
                events.medium,
                events.count,
                events.published_year,
            ]
        ).select_from(events_alias)
        return query
