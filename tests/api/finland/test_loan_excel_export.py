import io
import unittest
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from api.local_analytics_exporter import LocalAnalyticsExporter

LOAN_DB_FIXTURE = [
    {
        "identifier": "https://standardebooks.org/ebooks/charles-dickens/bleak-house",
        "identifier_type": "URI",
        "sort_title": "Bleak House",
        "sort_author": "Dickens, Charles",
        "publisher": "Standard Ebooks",
        "language": "eng",
        "genres": None,
        "contributors": ["Dickens, Charles (Author)"],
        "location": None,
        "library_name": "Open Access Library",
        "medium": "Book",
        "count": 5,
    },
    {
        "identifier": "https://standardebooks.org/ebooks/george-eliot/silas-marner",
        "identifier_type": "URI",
        "sort_title": "Silas Marner",
        "sort_author": "Eliot, George",
        "fiction": "fiction",
        "publisher": "Standard Ebooks",
        "language": "eng",
        "genres": ["Literary Fiction"],
        "contributors": ["Eliot, George (Author)", "Fake contributor (Author)"],
        "location": None,
        "library_name": "Open Access Library",
        "medium": "Book",
        "count": 8,
    },
    {
        "identifier": "https://standardebooks.org/ebooks/bertrand-russell/roads-to-freedom",
        "identifier_type": "URI",
        "sort_title": "Roads to Freedom",
        "sort_author": "Russell, Bertrand",
        "fiction": "non-fiction",
        "publisher": "Standard Ebooks",
        "language": "eng",
        "genres": ["Philosophy"],
        "contributors": ["Russell, Bertrand (Author)"],
        "location": None,
        "library_name": "Open Access Library",
        "medium": "Book",
        "count": 2,
    },
]


class TestExcelExport(unittest.TestCase):
    def test_export_excel(self):
        # Mock the database connection and its execute method
        mock_db = MagicMock()
        mock_db.execute.return_value = LOAN_DB_FIXTURE

        exporter = LocalAnalyticsExporter()

        # Patch the database connection and run the export method
        with patch.object(exporter, "analytics_query_loan_statistics") as mock_query:
            mock_query.return_value = "Mock SQL query"
            stream = exporter.export_excel(mock_db, "2023-01-01", "2023-12-31")

        # Load the stream into an openpyxl workbook
        bytes_in = io.BytesIO(stream)
        wb = load_workbook(bytes_in)
        sheet = wb.active

        # Validate the header row content
        expected_headers = (
            "Tekijä (aakkostus)",
            "Nimeke",
            "Fiktio",
            "Tunniste",
            "Tunnisteen tyyppi",
            "Kirjasto",
            "Sijainti",
            "Formaatti",
            "Kategoria(t)",
            "Kieli",
            "Kustantaja/Julkaisija",
            "Kaikki lainat",
            "Tekijä 1",
            "Tekijä 2",
        )
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        self.assertEqual(header_row, expected_headers, "Header row mismatch")

        # Validate the number of rows in the worksheet
        expected_row_count = len(LOAN_DB_FIXTURE)
        actual_row_count = sheet.max_row - 1  # Subtracting the header row
        self.assertEqual(actual_row_count, expected_row_count, "Row count mismatch")

        # Validate database query interaction
        mock_db.execute.assert_called_once_with("Mock SQL query")


if __name__ == "__main__":
    unittest.main()
